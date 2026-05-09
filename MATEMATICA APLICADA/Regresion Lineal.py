#-=========================================================DOCUMENTACION INTERNA=================================================================================================
# -- Objetivo: Implementar un programa que permita al usuario graficar puntos en un plano cartesiano y calcular su regresión lineal.
# -- Autor: Elmer Juarez / Correo de contacto: e24.elmerjuarez.juareza@suizoamaericano.edu.gt
# -- Descripción: El programa permite al usuario ingresar puntos en un plano cartesiano ya sea manualmente (click) o cargando un archivo Excel,
#                 calcular la regresión lineal de los puntos y mostrar la recta resultante en la interfaz gráfica.
# -- Lenguaje: Python
# -- Recursos: Python 3, librerías tkinter, scipy, openpyxl.
# -- Procesos: 
#       1. El usuario ingresa puntos haciendo click en el canvas o cargando un archivo Excel (.xlsx / .xls) desde A2:B2.
#       2. Los puntos se muestran en el plano cartesiano y en una tabla de coordenadas.
#       3. El programa calcula la regresión lineal usando scipy.stats.linregress.
#       4. La recta de regresión se dibuja en el plano cartesiano en color rojo.
#       5. La escala del plano se ajusta automáticamente la primera vez y se mantiene fija en actualizaciones posteriores.
# -- Historia: 
#       Fecha de creación: 24/04/2026
#       Fecha de modificación: 07/05/2026
# -- Ajustes pendientes: Mejorar la precisión de la conversión de coordenadas al hacer click, agregar opción para limpiar los puntos manualmente.
# -- Cambios Realizados: Implementación de carga de coordenadas desde archivo Excel, corrección de escala fija al actualizar la recta de regresión.
# ======================================================================================================================================================================================-->


from tkinter import *
from tkinter import ttk
from tkinter import messagebox as mb
from tkinter import filedialog
from scipy import stats
import openpyxl#Lineria que permite abrir el archivo de excel ¡¡Debe descargarla antes de utilizar el programa!!


class app(Tk): 
    def __init__(self):
        super().__init__()
        self.geometry("1050x600")
        
        # Canvas principal
        self.canva = Canvas(width=580, height=580, bg="lightblue")
        self.canva.place(x=10,y=10)
        
        # Frame para botones
        frame_botones = Frame(self)
        frame_botones.place(x=880, y=10)
        
        Button(frame_botones, text="Regresión Linear", command=self.regression, bg=("pink"), width=20, height=2).pack()
        Button(frame_botones, text="Abrir Excel", command=self.cargar_excel, bg=("lightgreen"), width=20, height=2).pack()
        
        # Evento click
        self.canva.bind("<Button-1>", self.posicionar)

        # Lista de coordenadas
        self.cords = []

        # Escala inicial
        self.escala = 6

        # Dibujar plano inicial
        self.crear_plano()
        
        self.uso = 0

    def cargar_excel(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")]
        )
        
        if not ruta:
            return
        
        try:
            wb = openpyxl.load_workbook(ruta)
            ws = wb.active
            
            nuevas = []
            fila = 2  # Empieza en A2, B2
            while True:
                val_x = ws[f"A{fila}"].value
                val_y = ws[f"B{fila}"].value
                
                if val_x is None and val_y is None:
                    break
                
                if val_x is not None and val_y is not None:
                    nuevas.append((float(val_x), float(val_y)))
                
                fila += 1
            
            if not nuevas:
                mb.showwarning("Aviso", "No se encontraron coordenadas válidas desde A2:B2.")
                return
            
            # Limpiar puntos anteriores y cargar los nuevos
            self.cords = []
            self.canva.delete("puntos")
            self.uso = 0  # Resetea escala al cargar nuevo Excel
            
            for par in nuevas:
                self.cords.append(par)
            
            self.mostrar_todos()
            self.regression()
                
        except Exception as e:
            mb.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def mostrar_todos(self):
        # Redibujar puntos
        w = self.canva.winfo_width() / 2
        h = self.canva.winfo_height() / 2
        
        for par in self.cords:
            px = par[0] * self.escala + w
            py = h - (par[1] * self.escala)
            self.canva.create_oval(px-3, py-3, px+3, py+3, fill="yellow", outline="", tags="puntos")
        
        # Actualizar tabla
        columnas = ["x", "y"]
        self.tree = ttk.Treeview(columns=columnas, show='headings', height=25)
        for column in columnas:
            self.tree.heading(column, text=column)
            self.tree.column(column, width=100, anchor='center')
        self.tree.place(x=650, y=30)
        for par in self.cords:
            self.tree.insert("", "end", values=(round(par[0], 2), round(par[1], 2)))

    def posicionar(self, event):
        self.cordenadas(event.x, event.y)
        self.canva.create_oval(event.x-3, event.y-3, event.x+3, event.y+3, fill="yellow", outline="", tags="puntos")
        if self.uso > 0:
            self.regression()
    
    def crear_plano(self): 
        self.canva.delete("grid")
        tam = 600
        origen = tam // 2
        escala = getattr(self, "escala", 6)
        espacio = escala * 5

        for i in range(0, tam, int(espacio)):
            self.canva.create_line(i, 0, i, tam, fill="gray", tags="grid")
            self.canva.create_line(0, i, tam, i, fill="gray", tags="grid")

        self.canva.create_line(origen, 0, origen, tam, width=2, tags="grid")
        self.canva.create_line(0, origen, tam, origen, width=2, tags="grid")

        for j in range(0, tam, int(espacio)): 
            self.canva.create_line(j, origen-5, j, origen+5, tags="grid")
            n = (j - origen) / escala
            if int(n) != 0:
                self.canva.create_text(j, origen+15, text=str(int(n)), tags="grid")
            self.canva.create_line(origen-5, j, origen+5, j, tags="grid")
            n = (origen - j) / escala
            if int(n) != 0:
                self.canva.create_text(origen-20, j, text=str(int(n)), tags="grid")

    def cordenadas(self, x:int, y:int):  
        w = self.canva.winfo_width() / 2
        h = self.canva.winfo_height() / 2
        escala = getattr(self, "escala", 6)
        x_cart = (x - w) / escala
        y_cart = (h - y) / escala
        par = (round(x_cart-1), round(y_cart+1))
        self.mostrar(par)
    
    def mostrar(self, par): 
        columnas = ["x", "y"]
        self.tree = ttk.Treeview(columns=columnas, show='headings', height=25)
        for column in columnas: 
            self.tree.heading(column, text=column)
            self.tree.column(column, width=100, anchor='center')
        self.tree.place(x=650, y=10)
        self.cords.append(par)
        for i in range(len(self.cords)): 
            self.tree.insert("", "end", values=self.cords[i])
    
    def regression(self): 
        self.uso += 1
        self.canva.delete("linea")
        x = []
        y = []
        for punto in self.cords:
            x.append(punto[0])
            y.append(punto[1])
        
        slope, intercept, r, p, std_err = stats.linregress(x, y) 

        max_x = max([abs(i) for i in x])
        max_y = max([abs(i) for i in y])
        max_val = max(max_x, max_y, 1)

        if self.uso == 1:  # Solo ajusta la escala la primera vez
            escala = (self.canva.winfo_width() / 2 - 20) / max_val
            self.escala = escala

        self.crear_plano()

        w = self.canva.winfo_width() / 2
        h = self.canva.winfo_height() / 2

        x1 = -max_val * 1.5
        x2 = max_val * 1.5
        y1 = slope * x1 + intercept
        y2 = slope * x2 + intercept

        px1 = x1 * self.escala + w
        py1 = h - (y1 * self.escala)
        px2 = x2 * self.escala + w
        py2 = h - (y2 * self.escala)

        self.canva.create_line(px1, py1, px2, py2, fill="red", width=2, tags="linea")

app().mainloop()